import os
from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy

# Chemin vers le dossier contenant les fichiers Excel
dossier_fichiers = './PROPOSITIONS'

# Crée un nouveau classeur Excel pour regrouper les fichiers
combined = Workbook()

# Supprime la feuille par défaut si elle existe
if 'Sheet' in combined.sheetnames:
    del combined['Sheet']

# Parcourir tous les fichiers dans le dossier
for fichier in os.listdir(dossier_fichiers):
    if fichier.endswith('.xlsx'):
        # Chemin complet vers le fichier Excel
        chemin_complet = os.path.join(dossier_fichiers, fichier)

        # Charger le fichier Excel
        wb = load_workbook(chemin_complet)

        # Copier chaque feuille dans le nouveau classeur
        for feuille in wb.sheetnames:
            feuille_source = wb[feuille]

            # Créer une nouvelle feuille avec les 4 premiers caractères du fichier
            feuille_nom = f"{fichier[:4]}_{feuille[:27]}"  # Limiter à 31 caractères (4 + 27 max)
            nouvelle_feuille = combined.create_sheet(title=feuille_nom)

            # Copier les données et les styles de chaque cellule
            for row in feuille_source.iter_rows():
                for cell in row:
                    nouvelle_cellule = nouvelle_feuille[cell.coordinate]
                    nouvelle_cellule.value = cell.value

                    # Copier les styles si la cellule en possède
                    if cell.has_style:
                        nouvelle_cellule.font = copy(cell.font)
                        nouvelle_cellule.border = copy(cell.border)
                        nouvelle_cellule.fill = copy(cell.fill)
                        nouvelle_cellule.number_format = copy(cell.number_format)
                        nouvelle_cellule.protection = copy(cell.protection)
                        nouvelle_cellule.alignment = copy(cell.alignment)

# Sauvegarder le fichier combiné
fichier_final = 'classeur_combiné_preservé.xlsx'
combined.save(fichier_final)

print(f"Les fichiers Excel ont été regroupés dans {fichier_final}, avec formatage préservé.")
